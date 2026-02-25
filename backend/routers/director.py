
from fastapi import APIRouter, Query
from app.services.director_service import DirectorService
from app.repositories.director_repository import DirectorRepository
from fastapi.responses import StreamingResponse
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from io import BytesIO
from datetime import datetime
from app.repositories.docente_repository import DocenteRepository
import mysql.connector

router = APIRouter(
	prefix="/director",
	tags=["Director"]
)

@router.get("/dashboard-metrics")
def get_director_dashboard_metrics(user_id: int = Query(..., description="ID del usuario director")):
	"""Obtiene las métricas del dashboard para un director específico"""
	return DirectorService.obtener_metricas_dashboard(user_id)


@router.get("/activities")
def get_activities_by_role(
    user_id: int = Query(..., description="ID del usuario"),
    role: str = Query(..., description="Rol del usuario: admin, director, docente")
):
    try:
        if role == "admin":
            return DirectorRepository.obtener_todas_actividades()
        elif role == "director":
            return DirectorRepository.obtener_actividades_director(user_id)
        elif role == "docente":
            return DocenteRepository.obtener_actividades_docente(user_id)
        else:
            return {"status": "error", "message": "Rol no válido"}
    except Exception as e:
        return {"status": "error", "message": str(e)}
    
@router.get("/reports/excel")
def download_director_reports_excel(
    user_id: int = Query(..., description="ID del usuario director"),
    fecha_inicio: str = None,
    fecha_final: str = None,
    estado: str = None,
    unidad: str = None
):
    """Descarga un Excel de actividades solo de las unidades del director"""
    try:
        actividades = DirectorRepository.obtener_actividades_director_para_reporte(user_id, fecha_inicio, fecha_final, estado, unidad)

        wb = Workbook()
        ws = wb.active
        ws.title = "Reporte"

        headers = [
            "Usuario",
            "Unidad",
            "Actividad",
            "Descripción",
            "Fecha envío",
            "Mes",
            "Estado"
        ]
        ws.append(headers)

        header_fill = PatternFill(fill_type="solid", fgColor="1F6FEB")
        header_font = Font(color="FFFFFF", bold=True)
        thin = Side(style="thin", color="D0D7DE")
        cell_border = Border(left=thin, right=thin, top=thin, bottom=thin)

        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = cell_border

        month_names = {
            1: "Enero", 2: "Febrero", 3: "Marzo", 4: "Abril", 5: "Mayo", 6: "Junio",
            7: "Julio", 8: "Agosto", 9: "Septiembre", 10: "Octubre", 11: "Noviembre", 12: "Diciembre"
        }

        for item in actividades:
            created_at = item.get("created_at")
            fecha_envio = created_at.strftime("%d-%m-%Y") if created_at else "-"
            mes = month_names.get(created_at.month, "-") if created_at else "-"

            ws.append([
                item.get("user_name") or "-",
                item.get("unit") or "-",
                item.get("code") or "-",
                item.get("description") or "-",
                fecha_envio,
                mes,
                item.get("state") or "-"
            ])

        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=7):
            for cell in row:
                cell.border = cell_border
                cell.alignment = Alignment(vertical="top", wrap_text=True)

        ws.freeze_panes = "A2"
        ws.auto_filter.ref = None

        column_widths = [24, 28, 24, 44, 14, 14, 18]
        for index, width in enumerate(column_widths, start=1):
            ws.column_dimensions[chr(64 + index)].width = width

        file_stream = BytesIO()
        wb.save(file_stream)
        file_stream.seek(0)

        filename = f"reporte_actividades_director_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"

        return StreamingResponse(
            file_stream,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f'attachment; filename="{filename}"'}
        )

    except Exception as e:
        print(f"❌ Error generando Excel de reportes director: {e}")
        return {
            "status": "error",
            "message": str(e)
        }